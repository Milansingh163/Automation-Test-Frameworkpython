import openpyxl


def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def getColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column

def readData(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)



# path  = ".\\TestData\\logindata.xlsx"
# workbook = openpyxl.load_workbook(path,'Sheet1')
# # sheet = workboo
# sheet = workbook.active
# rows = sheet.max_row
# cols = sheet.max_column
# # print(rows)
# # print(cols)

# for r in range(2,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(row=r,column=c).value,end="             ")
#     print()

# path2  = ".\\TestData\\writedata.xlsx"
# workbook = openpyxl.load_workbook(path2)
# sheet = workbook.active
# for r in range(1,5):
#     for c in range(1,4):
#         sheet.cell(row=r,column=c).value='milan'
#         print(sheet.cell(row=r,column=c).value,end='   ')
#     print()

# workbook.save(path2)


# def read_data(file,sheetname,row_number,col_number):
#     workbook = openpyxl.load_workbook(filename=file)
#     sheet = workbook[sheetname]
#     return sheet.cell(row=row_number,column=col_number)

# print(read_data("TestData/logindata.xlsx",'Sheet1',1,1))